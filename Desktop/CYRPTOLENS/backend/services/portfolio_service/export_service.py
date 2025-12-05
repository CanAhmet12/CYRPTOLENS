"""
Export Service for Portfolio.
Handles PDF, CSV, Excel export generation.
"""
from typing import List, Dict, Optional, Any
from decimal import Decimal
from datetime import datetime, date
from io import BytesIO
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ExportService:
    """Service for portfolio data export."""
    
    def generate_pdf_report(
        self,
        portfolio_data: Dict,
        period: str = 'ALL'
    ) -> BytesIO:
        """
        Generate PDF report.
        
        Args:
            portfolio_data: Portfolio data dict
            period: Period string (1M, 3M, 6M, 1Y, ALL)
        
        Returns:
            BytesIO object with PDF content
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#FF6B35'),
            spaceAfter=30
        )
        story.append(Paragraph('Portfolio Report', title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary section
        story.append(Paragraph('Portfolio Summary', styles['Heading2']))
        summary_data = [
            ['Total Value', f"${portfolio_data.get('total_value', 0):,.2f}"],
            ['Total Cost', f"${portfolio_data.get('total_cost', 0):,.2f}"],
            ['Total Profit/Loss', f"${portfolio_data.get('total_profit_loss', 0):,.2f}"],
            ['Profit/Loss %', f"{portfolio_data.get('total_profit_loss_percent', 0):.2f}%"],
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Holdings section
        if 'items' in portfolio_data:
            story.append(Paragraph('Holdings', styles['Heading2']))
            holdings_data = [['Coin', 'Amount', 'Buy Price', 'Current Price', 'Value', 'P/L', 'P/L %']]
            
            for item in portfolio_data['items']:
                holdings_data.append([
                    item.get('coin_symbol', ''),
                    f"{item.get('amount', 0):.8f}",
                    f"${item.get('buy_price', 0):,.2f}",
                    f"${item.get('current_price', 0):,.2f}",
                    f"${item.get('total_value', 0):,.2f}",
                    f"${item.get('profit_loss', 0):,.2f}",
                    f"{item.get('profit_loss_percent', 0):.2f}%"
                ])
            
            holdings_table = Table(holdings_data, colWidths=[0.8*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
            holdings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            story.append(holdings_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_csv_export(
        self,
        portfolio_data: Dict,
        export_type: str = 'full'
    ) -> bytes:
        """
        Generate CSV export.
        
        Args:
            portfolio_data: Portfolio data dict
            export_type: Type of export (transactions, tax, full)
        
        Returns:
            CSV string
        """
        output = []
        writer = csv.writer(output)
        
        if export_type == 'full' or export_type == 'portfolio':
            # Portfolio summary
            writer.writerow(['Portfolio Summary'])
            writer.writerow(['Total Value', f"${portfolio_data.get('total_value', 0):,.2f}"])
            writer.writerow(['Total Cost', f"${portfolio_data.get('total_cost', 0):,.2f}"])
            writer.writerow(['Total Profit/Loss', f"${portfolio_data.get('total_profit_loss', 0):,.2f}"])
            writer.writerow(['Profit/Loss %', f"{portfolio_data.get('total_profit_loss_percent', 0):.2f}%"])
            writer.writerow([])
            
            # Holdings
            writer.writerow(['Holdings'])
            writer.writerow(['Coin', 'Amount', 'Buy Price', 'Current Price', 'Value', 'P/L', 'P/L %'])
            for item in portfolio_data.get('items', []):
                writer.writerow([
                    item.get('coin_symbol', ''),
                    item.get('amount', 0),
                    item.get('buy_price', 0),
                    item.get('current_price', 0),
                    item.get('total_value', 0),
                    item.get('profit_loss', 0),
                    item.get('profit_loss_percent', 0)
                ])
        
        if export_type == 'transactions' or export_type == 'full':
            writer.writerow([])
            writer.writerow(['Transactions'])
            writer.writerow(['Date', 'Type', 'Coin', 'Amount', 'Price', 'Fee', 'Total Cost'])
            for tx in portfolio_data.get('transactions', []):
                writer.writerow([
                    tx.get('transaction_date', ''),
                    tx.get('transaction_type', ''),
                    tx.get('coin_symbol', ''),
                    tx.get('amount', 0),
                    tx.get('price', 0),
                    tx.get('fee', 0),
                    tx.get('total_cost', 0)
                ])
        
        csv_string = '\n'.join([''.join(row) for row in output])
        return csv_string.encode('utf-8')
    
    def generate_excel_export(
        self,
        portfolio_data: Dict
    ) -> BytesIO:
        """
        Generate Excel export.
        
        Args:
            portfolio_data: Portfolio data dict
        
        Returns:
            BytesIO object with Excel content
        """
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        ws_summary.append(['Portfolio Summary'])
        ws_summary.append(['Total Value', portfolio_data.get('total_value', 0)])
        ws_summary.append(['Total Cost', portfolio_data.get('total_cost', 0)])
        ws_summary.append(['Total Profit/Loss', portfolio_data.get('total_profit_loss', 0)])
        ws_summary.append(['Profit/Loss %', portfolio_data.get('total_profit_loss_percent', 0)])
        
        # Holdings sheet
        ws_holdings = wb.create_sheet('Holdings')
        ws_holdings.append(['Coin', 'Amount', 'Buy Price', 'Current Price', 'Value', 'P/L', 'P/L %'])
        for item in portfolio_data.get('items', []):
            ws_holdings.append([
                item.get('coin_symbol', ''),
                item.get('amount', 0),
                item.get('buy_price', 0),
                item.get('current_price', 0),
                item.get('total_value', 0),
                item.get('profit_loss', 0),
                item.get('profit_loss_percent', 0)
            ])
        
        # Format headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for ws in [ws_summary, ws_holdings]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

